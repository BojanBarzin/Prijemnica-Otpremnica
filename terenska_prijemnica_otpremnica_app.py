import base64
from datetime import date
from io import BytesIO
import os
import re
import zipfile

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import streamlit.components.v1 as components

st.set_page_config(page_title="Prijemnica / Otpremnica", layout="wide")

BRAND_YELLOW = "#FFD700"
GRAPHITE = "#111111"
LIGHT_GRAY = "#F5F5F5"

DATA_FILE = "data.xlsx"
LOCATIONS_FILE = "LocationsSPTS.csv"
TEMPLATE_XLSX = "Prijamnica-otpremnica-template.xlsx"
MAX_ITEMS = 5

COMMON_CLIENTS = [
    "Tendam",
    "Deichmann",
    "Takko",
    "Mercator-S",
    "H&M",
    "Metre Cash & Carry",
    "Ikea",
    "Decathlon",
    "Lidl",
]

MAGACINI = [
    "FSBG",
    "FS NIŠ",
    "FSNS",
    "FS ZAGREB",
    "FS BANJA LUKA",
    "FS SARAJEVO",
]

TEHNICARI = [
    "Adel Imeri",
    "Andrej Prole",
    "Bojan Barzin",
    "Damir Ulovec",
    "Dejan Todorovic",
    "Goran Krstic",
    "Ivan Petracic",
    "Jovan Ilic",
    "Kresimir Vukman",
    "Lauro Cuha",
    "Marko Madjar",
    "Miroslav Hicil",
    "Mladjen Stojakovic",
    "Zvonko Jovic",
]

PROJECT_LABELS = {
    "107": "Tendam",
    "108": "Deichmann",
    "109": "Takko",
    "112": "Mercator-S",
    "115": "H&M",
    "118": "Metre Cash & Carry",
    "119": "Ikea",
    "123": "Decathlon",
    "193": "Lidl",
}

# =========================
# STYLE
# =========================
def get_base64(path: str) -> str:
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except Exception:
        return ""

bg_logo = get_base64("assets/fs_logo_white.png")


def apply_style():
    st.markdown(
        f"""
        <style>
        html, body, [class*="css"] {{
            font-family: 'Nunito Sans', 'Segoe UI', sans-serif;
        }}
        .stApp {{
            background-color: {LIGHT_GRAY};
            overflow-x: hidden;
        }}
        .stApp::before {{
            content: "";
            position: fixed;
            inset: -300px;
            background-image: url("data:image/png;base64,{bg_logo}");
            background-size: 340px;
            background-repeat: repeat;
            background-position: 0 0;
            opacity: 0.045;
            transform: rotate(-18deg);
            z-index: 0;
            pointer-events: none;
            animation: bgMove 55s linear infinite;
        }}
        @keyframes bgMove {{
            from {{ background-position: 0 0; }}
            to {{ background-position: 900px 900px; }}
        }}
        .block-container {{
            position: relative;
            z-index: 1;
            max-width: 1280px;
            padding-top: 1.4rem;
        }}
        .brand-header {{
            background: rgba(17,17,17,0.96);
            padding: 20px 26px;
            border-radius: 18px;
            margin-bottom: 18px;
            border-left: 10px solid {BRAND_YELLOW};
            box-shadow: 0 8px 24px rgba(0,0,0,0.18);
        }}
        .brand-title {{
            color: white;
            font-size: 31px;
            font-weight: 900;
            margin: 0;
        }}
        .brand-subtitle {{
            color: #d9d9d9;
            font-size: 15px;
            margin-top: 4px;
        }}
        .paper {{
            background: white;
            border: 1px solid #222;
            padding: 18px 20px;
            box-shadow: 0 8px 22px rgba(0,0,0,0.10);
            margin-bottom: 18px;
        }}
        .doc-top {{
            display: grid;
            grid-template-columns: 1fr 1fr 1fr;
            align-items: start;
            margin-bottom: 8px;
        }}
        .company {{
            text-align: center;
            font-size: 13px;
            font-weight: 600;
        }}
        .doc-title {{
            font-size: 17px;
            font-weight: 800;
            text-align: left;
            padding-top: 6px;
        }}
        .stTextInput label, .stDateInput label, div[data-baseweb="select"] label {{
            font-weight: 700 !important;
            color: #111 !important;
        }}
        .stTextInput input,
        .stDateInput input,
        .stNumberInput input {{
            background-color: white !important;
            color: black !important;
            border: 1px solid #222 !important;
            border-radius: 0 !important;
            min-height: 34px !important;
        }}
        div[data-baseweb="select"] > div {{
            background-color: white !important;
            color: black !important;
            border: 1px solid #222 !important;
            border-radius: 0 !important;
            min-height: 34px !important;
        }}
        div[data-baseweb="select"] svg {{ display: none !important; }}
        div[data-baseweb="select"] [aria-label="open"] {{ display: none !important; }}
        .section-line {{ border-top: 1px solid #222; margin: 18px 0 14px 0; }}
        .item-header {{
            display: grid;
            grid-template-columns: 50px 1.4fr 1.2fr 1.3fr 1.3fr 1.2fr;
            border-top: 1px solid #222;
            border-left: 1px solid #222;
            font-size: 13px;
            font-weight: 800;
            text-align: center;
            margin-top: 10px;
        }}
        .item-header div {{
            border-right: 1px solid #222;
            border-bottom: 1px solid #222;
            padding: 7px 4px;
        }}
        .sign-row {{
            display: grid;
            grid-template-columns: 1fr 1fr 1fr;
            gap: 18px;
            margin-top: 16px;
        }}
        .sign-box {{
            border-top: 1px solid #222;
            text-align: center;
            padding-top: 6px;
            font-size: 13px;
            font-weight: 700;
            min-height: 44px;
        }}
        div.stButton > button,
        button[kind="secondary"] {{
            background: {GRAPHITE} !important;
            color: white !important;
            border: 1px solid {GRAPHITE} !important;
            border-radius: 10px !important;
            font-weight: 700 !important;
        }}
        div.stButton > button:hover,
        button[kind="secondary"]:hover {{
            background: black !important;
            color: {BRAND_YELLOW} !important;
            border: 1px solid {BRAND_YELLOW} !important;
        }}
        .stDownloadButton > button {{
            background: {BRAND_YELLOW} !important;
            color: black !important;
            border-radius: 10px !important;
            font-weight: 800 !important;
            border: none !important;
        }}
        @media print {{
            .no-print, header, footer, [data-testid="stToolbar"] {{ display: none !important; }}
            .paper {{ box-shadow: none; margin: 0; page-break-inside: avoid; }}
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

apply_style()

# Header
col_logo, col_title = st.columns([1, 5])
with col_logo:
    try:
        st.image("assets/fs_logo.png", width=120)
    except Exception:
        pass
with col_title:
    st.markdown(
        """
        <div class="brand-header">
            <div class="brand-title">Prijemnica / Otpremnica sa terena</div>
            <div class="brand-subtitle">Prediktivno popunjavanje uređaja i lokacije iz CMDB i SPTS baze</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

# =========================
# DATA LOAD
# =========================
@st.cache_data
def load_cmdb():
    try:
        data = pd.read_excel(DATA_FILE, dtype=str).fillna("")
    except Exception:
        data = pd.DataFrame()
    data.columns = [str(c).strip() for c in data.columns]
    return data


@st.cache_data
def load_locations():
    candidates = [
        LOCATIONS_FILE,
        "LocationsSPTS(1).csv",
        "locations.csv",
    ]
    for file_name in candidates:
        if os.path.exists(file_name):
            try:
                data = pd.read_csv(file_name, dtype=str).fillna("")
                data.columns = [str(c).strip() for c in data.columns]
                return data
            except Exception:
                continue
    return pd.DataFrame()

cmdb = load_cmdb()
locations = load_locations()

if cmdb.empty:
    st.error(f"Nije pronađen ili je prazan fajl: {DATA_FILE}")
    st.stop()

if locations.empty:
    st.warning(f"Nije pronađen ili je prazan fajl: {LOCATIONS_FILE}. Lokacije neće biti dostupne za predloge.")

# =========================
# HELPERS
# =========================
def col_value(row, *names):
    for name in names:
        if name in row.index:
            val = row.get(name, "")
            if pd.notna(val):
                return str(val).strip()
    return ""


def get_options_from_df(df: pd.DataFrame, col: str, fallback=None):
    fallback = fallback or []
    if col not in df.columns:
        return fallback
    values = (
        df[col]
        .astype(str)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .drop_duplicates()
        .tolist()
    )
    return sorted(list(dict.fromkeys(values + fallback)))


def map_project_to_name(project_value: str) -> str:
    value = str(project_value).strip()
    return PROJECT_LABELS.get(value, value)


def object_options():
    if locations.empty or "Name" not in locations.columns:
        return []
    return get_options_from_df(locations, "Name")


def object_row_by_name(name: str):
    if locations.empty or "Name" not in locations.columns or not name:
        return None
    name_norm = str(name).strip().lower()
    exact = locations[locations["Name"].astype(str).str.strip().str.lower() == name_norm]
    if not exact.empty:
        return exact.iloc[0]
    starts = locations[locations["Name"].astype(str).str.strip().str.lower().str.startswith(name_norm)]
    if len(starts) == 1:
        return starts.iloc[0]
    return None


def infer_city(address: str) -> str:
    if not address:
        return ""
    known = [
        "Beograd", "Novi Sad", "Niš", "Nis", "Kragujevac", "Subotica", "Zrenjanin", "Leskovac",
        "Čačak", "Cacak", "Kruševac", "Krusevac", "Pančevo", "Pancevo", "Kraljevo",
        "Užice", "Uzice", "Valjevo", "Sombor", "Kikinda", "Vršac", "Vrsac", "Šabac", "Sabac",
        "Sremska Mitrovica", "Loznica", "Jagodina", "Paraćin", "Paracin", "Pirot", "Zaječar", "Zajecar",
    ]
    a = address.lower()
    for city in known:
        if city.lower() in a:
            return "Niš" if city == "Nis" else city
    return ""


def device_options(col: str):
    return get_options_from_df(cmdb, col) if col in cmdb.columns else []


def find_device(inv: str, serial: str, sp: str):
    filters = []
    if inv and "inventory_number" in cmdb.columns:
        filters.append(("inventory_number", inv))
    if serial and "serial_number" in cmdb.columns:
        filters.append(("serial_number", serial))
    if sp and "sp_inventory_number" in cmdb.columns:
        filters.append(("sp_inventory_number", sp))

    for col, value in filters:
        exact = cmdb[cmdb[col].astype(str).str.strip().str.lower() == value.strip().lower()]
        if not exact.empty:
            return exact.iloc[0]

    for col, value in filters:
        contains = cmdb[cmdb[col].astype(str).str.contains(re.escape(value), case=False, na=False)]
        if len(contains) == 1:
            return contains.iloc[0]

    return None


def smart_select(label: str, options, key: str):
    """Searchable field sa slobodnim unosom kada Streamlit podržava accept_new_options."""
    if key not in st.session_state:
        st.session_state[key] = ""
    try:
        current = st.session_state.get(key) or None
        index = None
        if current in options:
            index = options.index(current)
        return st.selectbox(
            label,
            options=options,
            index=index,
            placeholder="",
            key=key,
            accept_new_options=True,
        ) or ""
    except TypeError:
        return st.text_input(label, key=key)


def set_default(key, value):
    if value and not st.session_state.get(key):
        st.session_state[key] = value


def date_to_str(value):
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")
    return str(value or "")

# =========================
# STATE INIT
# =========================
def init_defaults():
    defaults = {
        "pri_razduzio": "",
        "pri_broj": "",
        "pri_datum": date.today(),
        "pri_u_magacin": "",
        "pri_objekat": "",
        "pri_adresa": "",
        "pri_mesto": "",
        "pri_predao": "",
        "pri_zaduzio": "",
        "pri_zaprimio": "",
        "otp_broj": "",
        "otp_datum": date.today(),
        "otp_iz_magacina": "",
        "otp_zaduzio": "",
        "otp_objekat": "",
        "otp_adresa": "",
        "otp_mesto": "",
        "otp_otpremio": "",
        "otp_primio": "",
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

    for side in ["pri", "otp"]:
        for i in range(MAX_ITEMS):
            for field in ["naziv", "model", "inv", "sn", "sp"]:
                key = f"{side}_{field}_{i}"
                if key not in st.session_state:
                    st.session_state[key] = ""


init_defaults()

# =========================
# AUTO-FILL BEFORE UI
# =========================
def autofill_device_side(side: str):
    for i in range(MAX_ITEMS):
        matched = find_device(
            st.session_state.get(f"{side}_inv_{i}", ""),
            st.session_state.get(f"{side}_sn_{i}", ""),
            st.session_state.get(f"{side}_sp_{i}", ""),
        )
        if matched is not None:
            set_default(f"{side}_naziv_{i}", col_value(matched, "name", "Name"))
            set_default(f"{side}_model_{i}", col_value(matched, "model", "Model"))
            set_default(f"{side}_inv_{i}", col_value(matched, "inventory_number", "InventoryNumber"))
            set_default(f"{side}_sn_{i}", col_value(matched, "serial_number", "SerialNumber"))
            set_default(f"{side}_sp_{i}", col_value(matched, "sp_inventory_number", "SPInventoryNumber"))


autofill_device_side("pri")
autofill_device_side("otp")

# Auto-fill prijemnica object
obj_row = object_row_by_name(st.session_state.get("pri_objekat", ""))
if obj_row is not None:
    address = col_value(obj_row, "Address", "address", "Adress", "adres", "adresa")
    project_name = map_project_to_name(col_value(obj_row, "Project", "project"))
    set_default("pri_adresa", address)
    if not st.session_state.get("pri_mesto"):
        st.session_state["pri_mesto"] = infer_city(address)
    if project_name and not st.session_state.get("pri_razduzio"):
        st.session_state["pri_razduzio"] = project_name

# Auto-fill otpremnica object
obj_row_otp = object_row_by_name(st.session_state.get("otp_objekat", ""))
if obj_row_otp is not None:
    address = col_value(obj_row_otp, "Address", "address", "Adress", "adres", "adresa")
    set_default("otp_adresa", address)
    if not st.session_state.get("otp_mesto"):
        st.session_state["otp_mesto"] = infer_city(address)

# Mirror prijemnica -> otpremnica for document data only, NOT devices
mirror_map = {
    "otp_broj": "pri_broj",
    "otp_datum": "pri_datum",
    "otp_iz_magacina": "pri_razduzio",
    "otp_zaduzio": "pri_u_magacin",
    "otp_objekat": "pri_objekat",
    "otp_adresa": "pri_adresa",
    "otp_mesto": "pri_mesto",
    "otp_otpremio": "pri_predao",
}
for dest, src in mirror_map.items():
    if not st.session_state.get(dest) and st.session_state.get(src):
        st.session_state[dest] = st.session_state[src]

name_opts = device_options("name")
model_opts = device_options("model")
inv_opts = device_options("inventory_number")
serial_opts = device_options("serial_number")
sp_opts = device_options("sp_inventory_number")
obj_opts = object_options()
client_opts = COMMON_CLIENTS
magacin_opts = MAGACINI
tehnicar_opts = TEHNICARI

# =========================
# DOCUMENT UI
# =========================
st.markdown('<div class="paper">', unsafe_allow_html=True)

st.markdown(
    """
    <div class="doc-top">
        <div></div>
        <div class="company">Fiscal Solutions d.o.o. &nbsp; Temerinska 102, 21000 Novi Sad</div>
        <div></div>
    </div>
    """,
    unsafe_allow_html=True,
)

c_title, c_broj, c_datum = st.columns([1.4, 1, 1])
with c_title:
    st.markdown('<div class="doc-title">PRIJEMNICA BR.</div>', unsafe_allow_html=True)
with c_broj:
    st.text_input("Broj prijemnice", key="pri_broj")
with c_datum:
    st.date_input("Datum", key="pri_datum")

c1, c2 = st.columns(2)
with c1:
    smart_select("UREĐAJ RAZDUŽIO (ime i prezime / naziv firme)", client_opts, "pri_razduzio")
with c2:
    smart_select("U magacin / Ime i prezime", magacin_opts, "pri_u_magacin")

c1, c2, c3 = st.columns([1.4, 1.4, 1])
with c1:
    smart_select("Objekat", obj_opts, "pri_objekat")
with c2:
    st.text_input("Adresa", key="pri_adresa")
with c3:
    st.text_input("Mesto", key="pri_mesto")

st.markdown('<div class="item-header"><div>BR</div><div>NAZIV</div><div>MODEL</div><div>INV</div><div>SN</div><div>SP/FS</div></div>', unsafe_allow_html=True)
for i in range(MAX_ITEMS):
    ci, cn, cm, cinv, csn, csp = st.columns([0.35, 1.4, 1.2, 1.3, 1.3, 1.2])
    with ci:
        st.text_input("", value=str(i + 1), disabled=True, label_visibility="collapsed", key=f"pri_br_disabled_{i}")
    with cn:
        smart_select("Naziv" if i == 0 else "", name_opts, f"pri_naziv_{i}")
    with cm:
        smart_select("Model" if i == 0 else "", model_opts, f"pri_model_{i}")
    with cinv:
        smart_select("Inventarni broj" if i == 0 else "", inv_opts, f"pri_inv_{i}")
    with csn:
        smart_select("Serijski broj" if i == 0 else "", serial_opts, f"pri_sn_{i}")
    with csp:
        smart_select("SP/FS broj" if i == 0 else "", sp_opts, f"pri_sp_{i}")

s1, s2, s3 = st.columns(3)
with s1:
    smart_select("Uređaj predao", tehnicar_opts, "pri_predao")
with s2:
    smart_select("Uređaj zadužio", tehnicar_opts, "pri_zaduzio")
with s3:
    smart_select("Uređaj zaprimio", tehnicar_opts, "pri_zaprimio")

st.markdown('<div class="section-line"></div>', unsafe_allow_html=True)

st.markdown(
    """
    <div class="doc-top">
        <div></div>
        <div class="company">Fiscal Solutions d.o.o. &nbsp; Temerinska 102, 21000 Novi Sad</div>
        <div></div>
    </div>
    """,
    unsafe_allow_html=True,
)

c_title, c_broj, c_datum = st.columns([1.4, 1, 1])
with c_title:
    st.markdown('<div class="doc-title">OTPREMNICA BR.</div>', unsafe_allow_html=True)
with c_broj:
    st.text_input("Broj otpremnice", key="otp_broj")
with c_datum:
    st.date_input("Datum ", key="otp_datum")

c1, c2 = st.columns(2)
with c1:
    st.text_input("Iz magacina / Ime i prezime", key="otp_iz_magacina")
with c2:
    st.text_input("Uređaj zadužio / Ime i prezime", key="otp_zaduzio")

c1, c2, c3 = st.columns([1.4, 1.4, 1])
with c1:
    smart_select("Objekat ", obj_opts, "otp_objekat")
with c2:
    st.text_input("Adresa ", key="otp_adresa")
with c3:
    st.text_input("Mesto ", key="otp_mesto")

st.markdown('<div class="item-header"><div>BR</div><div>NAZIV</div><div>MODEL</div><div>INV</div><div>SN</div><div>SP/FS</div></div>', unsafe_allow_html=True)
for i in range(MAX_ITEMS):
    ci, cn, cm, cinv, csn, csp = st.columns([0.35, 1.4, 1.2, 1.3, 1.3, 1.2])
    with ci:
        st.text_input("", value=str(i + 1), disabled=True, label_visibility="collapsed", key=f"otp_br_disabled_{i}")
    with cn:
        smart_select("Naziv " if i == 0 else "", name_opts, f"otp_naziv_{i}")
    with cm:
        smart_select("Model " if i == 0 else "", model_opts, f"otp_model_{i}")
    with cinv:
        smart_select("Inventarni broj " if i == 0 else "", inv_opts, f"otp_inv_{i}")
    with csn:
        smart_select("Serijski broj " if i == 0 else "", serial_opts, f"otp_sn_{i}")
    with csp:
        smart_select("SP/FS broj " if i == 0 else "", sp_opts, f"otp_sp_{i}")

s1, s2, s3 = st.columns(3)
with s1:
    smart_select("Uređaj otpremio", tehnicar_opts, "otp_otpremio")
with s2:
    smart_select("Uređaj zadužio ", tehnicar_opts, "otp_zaduzio_bottom")
with s3:
    smart_select("Uređaj primio", tehnicar_opts, "otp_primio")

st.markdown('</div>', unsafe_allow_html=True)

# =========================
# EXCEL EXPORT
# =========================
def write_cell(ws, cell, value):
    for merged_range in ws.merged_cells.ranges:
        if cell in merged_range:
            coord = merged_range.start_cell.coordinate
            ws[coord] = value
            ws[coord].alignment = Alignment(horizontal="center", vertical="center")
            return
    ws[cell] = value
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")


def is_valid_xlsx(path: str) -> bool:
    if not os.path.exists(path):
        return False
    try:
        with zipfile.ZipFile(path, "r") as zf:
            return "xl/workbook.xml" in zf.namelist()
    except Exception:
        return False


def create_fallback_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Interna prijemnica"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    bold = Font(bold=True)

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = [6, 28, 22, 22, 22, 20, 18][col - 1]

    def merge_write(cell_range, value, bold_font=False, fill=False):
        ws.merge_cells(cell_range)
        start = cell_range.split(":")[0]
        ws[start] = value
        ws[start].alignment = Alignment(horizontal="center", vertical="center")
        ws[start].font = bold if bold_font else Font()
        if fill:
            ws[start].fill = header_fill
        for row in ws[cell_range]:
            for c in row:
                c.border = border

    def row_header(row):
        headers = ["BR", "NAZIV", "MODEL", "INV", "SN", "SP/FS", "NAPOMENA"]
        for idx, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=idx)
            c.value = h
            c.font = bold
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

    merge_write("A1:G1", "Fiscal Solutions d.o.o.   Temerinska 102, 21000 Novi Sad", True)
    merge_write("A3:C3", "PRIJEMNICA BR.", True)
    merge_write("D3:E3", "")
    merge_write("F3:G3", "Datum")
    merge_write("A6:C6", "UREĐAJ RAZDUŽIO (ime i prezime / naziv firme)", True, True)
    merge_write("D6:G6", "U magacin / Ime i prezime", True, True)
    merge_write("A8:B8", "Objekat", True, True)
    merge_write("C8:E8", "Adresa", True, True)
    merge_write("F8:G8", "Mesto", True, True)
    row_header(11)
    for r in range(12, 17):
        for c in range(1, 8):
            ws.cell(r, c).border = border
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
    merge_write("A20:B20", "Uređaj predao", True, True)
    merge_write("C20:E20", "Uređaj zadužio", True, True)
    merge_write("F20:G20", "Uređaj zaprimio", True, True)

    merge_write("A24:G24", "Fiscal Solutions d.o.o.   Temerinska 102, 21000 Novi Sad", True)
    merge_write("A26:C26", "OTPREMNICA BR.", True)
    merge_write("D26:E26", "")
    merge_write("F26:G26", "Datum")
    merge_write("A29:C29", "Iz magacina / Ime i prezime", True, True)
    merge_write("D29:G29", "Uređaj zadužio / Ime i prezime", True, True)
    merge_write("A31:B31", "Objekat", True, True)
    merge_write("C31:E31", "Adresa", True, True)
    merge_write("F31:G31", "Mesto", True, True)
    row_header(34)
    for r in range(35, 40):
        for c in range(1, 8):
            ws.cell(r, c).border = border
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
    merge_write("A43:B43", "Uređaj otpremio", True, True)
    merge_write("C43:E43", "Uređaj zadužio", True, True)
    merge_write("F43:G43", "Uređaj primio", True, True)

    return wb


def load_template_or_fallback():
    candidates = [
        TEMPLATE_XLSX,
        "Prijamnica-otpremnica-template.xlsx",
        "/mnt/data/Prijamnica-otpremnica-template.xlsx",
    ]
    for path in candidates:
        if is_valid_xlsx(path):
            try:
                return load_workbook(path)
            except Exception:
                pass
    return create_fallback_workbook()


def fill_template() -> bytes:
    wb = load_template_or_fallback()
    ws = wb["Interna prijemnica"] if "Interna prijemnica" in wb.sheetnames else wb.active

    # Prijemnica header - original template layout
    write_cell(ws, "D3", st.session_state.get("pri_broj", ""))
    write_cell(ws, "E4", date_to_str(st.session_state.get("pri_datum", date.today())))
    write_cell(ws, "B7", st.session_state.get("pri_razduzio", ""))
    write_cell(ws, "E7", st.session_state.get("pri_u_magacin", ""))
    write_cell(ws, "C8", st.session_state.get("pri_objekat", ""))
    write_cell(ws, "C9", st.session_state.get("pri_adresa", ""))
    write_cell(ws, "C10", st.session_state.get("pri_mesto", ""))

    start_row = 13
    for i in range(MAX_ITEMS):
        r = start_row + i
        write_cell(ws, f"A{r}", i + 1)
        write_cell(ws, f"B{r}", st.session_state.get(f"pri_naziv_{i}", ""))
        write_cell(ws, f"C{r}", st.session_state.get(f"pri_model_{i}", ""))
        write_cell(ws, f"D{r}", st.session_state.get(f"pri_inv_{i}", ""))
        write_cell(ws, f"E{r}", st.session_state.get(f"pri_sn_{i}", ""))
        write_cell(ws, f"F{r}", st.session_state.get(f"pri_sp_{i}", ""))

    write_cell(ws, "B20", st.session_state.get("pri_predao", ""))
    write_cell(ws, "D20", st.session_state.get("pri_zaduzio", ""))
    write_cell(ws, "E20", st.session_state.get("pri_zaprimio", ""))

    # Otpremnica header - original template layout
    write_cell(ws, "D26", st.session_state.get("otp_broj", ""))
    write_cell(ws, "E27", date_to_str(st.session_state.get("otp_datum", date.today())))
    write_cell(ws, "B30", st.session_state.get("otp_iz_magacina", ""))
    write_cell(ws, "D30", st.session_state.get("otp_zaduzio", ""))
    write_cell(ws, "E31", st.session_state.get("otp_objekat", ""))
    write_cell(ws, "E32", st.session_state.get("otp_adresa", ""))
    write_cell(ws, "E33", st.session_state.get("otp_mesto", ""))

    start_row = 36
    for i in range(MAX_ITEMS):
        r = start_row + i
        write_cell(ws, f"A{r}", i + 1)
        write_cell(ws, f"B{r}", st.session_state.get(f"otp_naziv_{i}", ""))
        write_cell(ws, f"C{r}", st.session_state.get(f"otp_model_{i}", ""))
        write_cell(ws, f"D{r}", st.session_state.get(f"otp_inv_{i}", ""))
        write_cell(ws, f"E{r}", st.session_state.get(f"otp_sn_{i}", ""))
        write_cell(ws, f"F{r}", st.session_state.get(f"otp_sp_{i}", ""))

    write_cell(ws, "B43", st.session_state.get("otp_otpremio", ""))
    write_cell(ws, "D43", st.session_state.get("otp_zaduzio_bottom", ""))
    write_cell(ws, "E43", st.session_state.get("otp_primio", ""))

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

# =========================
# PRINT HTML SAME LAYOUT
# =========================
def esc(x):
    return str(x or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def item_rows_html(side: str):
    rows = ""
    for i in range(MAX_ITEMS):
        rows += f"""
        <tr>
            <td>{i + 1}</td>
            <td>{esc(st.session_state.get(f'{side}_naziv_{i}', ''))}</td>
            <td>{esc(st.session_state.get(f'{side}_model_{i}', ''))}</td>
            <td>{esc(st.session_state.get(f'{side}_inv_{i}', ''))}</td>
            <td>{esc(st.session_state.get(f'{side}_sn_{i}', ''))}</td>
            <td>{esc(st.session_state.get(f'{side}_sp_{i}', ''))}</td>
            <td></td>
        </tr>
        """
    return rows


def build_print_html() -> str:
    return f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 portrait; margin: 8mm; }}
            body {{ font-family: Arial, sans-serif; font-size: 11px; color: #000; }}
            button {{ margin-bottom: 10px; padding: 8px 14px; background:#111; color:white; border:0; border-radius:6px; font-weight:bold; }}
            .sheet {{ border:1px solid #000; padding:10px; margin-bottom:12px; }}
            .company {{ text-align:center; font-weight:bold; margin-bottom:8px; }}
            .title-line {{ display:grid; grid-template-columns:1fr 1fr 1fr; gap:8px; align-items:center; margin-bottom:8px; }}
            .title {{ font-size:16px; font-weight:bold; }}
            table {{ width:100%; border-collapse:collapse; margin-top:6px; }}
            td, th {{ border:1px solid #000; padding:5px; text-align:center; height:20px; }}
            th {{ font-weight:bold; background:#f2f2f2; }}
            .sign {{ display:grid; grid-template-columns:1fr 1fr 1fr; gap:22px; margin-top:36px; }}
            .sig {{ border-top:1px solid #000; text-align:center; padding-top:6px; font-weight:bold; min-height:32px; }}
            @media print {{ button {{ display:none; }} .sheet {{ page-break-inside: avoid; }} }}
        </style>
    </head>
    <body>
        <button onclick="window.print()">Print</button>

        <div class="sheet">
            <div class="company">Fiscal Solutions d.o.o. &nbsp; Temerinska 102, 21000 Novi Sad</div>
            <div class="title-line">
                <div class="title">PRIJEMNICA BR.</div>
                <div>{esc(st.session_state.get('pri_broj'))}</div>
                <div>Datum: {esc(date_to_str(st.session_state.get('pri_datum')))}</div>
            </div>
            <table>
                <tr><th>UREĐAJ RAZDUŽIO (ime i prezime / naziv firme)</th><th>U magacin / Ime i prezime</th></tr>
                <tr><td>{esc(st.session_state.get('pri_razduzio'))}</td><td>{esc(st.session_state.get('pri_u_magacin'))}</td></tr>
                <tr><th>Objekat</th><th>Adresa</th><th>Mesto</th></tr>
                <tr><td>{esc(st.session_state.get('pri_objekat'))}</td><td>{esc(st.session_state.get('pri_adresa'))}</td><td>{esc(st.session_state.get('pri_mesto'))}</td></tr>
            </table>
            <table>
                <tr><th>BR</th><th>NAZIV</th><th>MODEL</th><th>INV</th><th>SN</th><th>SP/FS</th><th>NAPOMENA</th></tr>
                {item_rows_html('pri')}
            </table>
            <div class="sign">
                <div class="sig">Uređaj predao<br>{esc(st.session_state.get('pri_predao'))}</div>
                <div class="sig">Uređaj zadužio<br>{esc(st.session_state.get('pri_zaduzio'))}</div>
                <div class="sig">Uređaj zaprimio<br>{esc(st.session_state.get('pri_zaprimio'))}</div>
            </div>
        </div>

        <div class="sheet">
            <div class="company">Fiscal Solutions d.o.o. &nbsp; Temerinska 102, 21000 Novi Sad</div>
            <div class="title-line">
                <div class="title">OTPREMNICA BR.</div>
                <div>{esc(st.session_state.get('otp_broj'))}</div>
                <div>Datum: {esc(date_to_str(st.session_state.get('otp_datum')))}</div>
            </div>
            <table>
                <tr><th>Iz magacina / Ime i prezime</th><th>Uređaj zadužio / Ime i prezime</th></tr>
                <tr><td>{esc(st.session_state.get('otp_iz_magacina'))}</td><td>{esc(st.session_state.get('otp_zaduzio'))}</td></tr>
                <tr><th>Objekat</th><th>Adresa</th><th>Mesto</th></tr>
                <tr><td>{esc(st.session_state.get('otp_objekat'))}</td><td>{esc(st.session_state.get('otp_adresa'))}</td><td>{esc(st.session_state.get('otp_mesto'))}</td></tr>
            </table>
            <table>
                <tr><th>BR</th><th>NAZIV</th><th>MODEL</th><th>INV</th><th>SN</th><th>SP/FS</th><th>NAPOMENA</th></tr>
                {item_rows_html('otp')}
            </table>
            <div class="sign">
                <div class="sig">Uređaj otpremio<br>{esc(st.session_state.get('otp_otpremio'))}</div>
                <div class="sig">Uređaj zadužio<br>{esc(st.session_state.get('otp_zaduzio_bottom'))}</div>
                <div class="sig">Uređaj primio<br>{esc(st.session_state.get('otp_primio'))}</div>
            </div>
        </div>
    </body>
    </html>
    """

# =========================
# ACTIONS
# =========================
st.markdown('<div class="no-print">', unsafe_allow_html=True)
col_a, col_b, col_c = st.columns([1, 1, 1])
with col_a:
    try:
        excel_bytes = fill_template()
        st.download_button(
            "Preuzmi Excel prijemnicu / otpremnicu",
            data=excel_bytes,
            file_name="prijemnica_otpremnica_teren.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.error(f"Ne mogu da napravim Excel. Detalj: {e}")
with col_b:
    if st.button("Prikaži / Print dokument"):
        components.html(build_print_html(), height=1300, scrolling=True)
with col_c:
    if st.button("Osveži automatsko popunjavanje"):
        st.rerun()
st.markdown('</div>', unsafe_allow_html=True)
